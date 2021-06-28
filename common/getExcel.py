from openpyxl import load_workbook
from common.getPath import *
from common.readConfig import ReadConfig
import pandas as pd

class DoExcel:
    @classmethod
    def read_excel(cls):
        '''获取excel表格'''
        # sort = pd.read_excel(file_path, sheet_name='init').values[0,0]
        # print(type(sort))
        wb = load_workbook(file_path)
        button = eval(ReadConfig.read_config('MODULE','mini_App'))
        test_list = []
        for key in button:
            sheet = wb[key]
            row = sheet.max_row
            if button[key] == "all":
                for x in range(2,row+1):
                    test_dict = {}
                    test_dict['id'] = sheet.cell(x,1).value
                    test_dict['module'] = sheet.cell(x, 2).value
                    test_dict['title'] = sheet.cell(x, 3).value
                    test_dict['method'] = sheet.cell(x, 4).value
                    test_dict['url'] = sheet.cell(x, 5).value
                    # if sheet.cell(x,6).value.find('${sort}') !=-1:
                    #     test_dict['data'] = sheet.cell(x,6).value.replace('${sort}',str(sort))
                    #     sort += 1
                    # else:
                    test_dict['data'] = sheet.cell(x,6).value
                    test_dict['expected'] = sheet.cell(x, 7).value
                    test_dict['sheet_name'] = key
                    test_list.append(test_dict)
                    '''将遍历后的字典数据都追加在test_list空表内'''
                    # cls.update_data(sheetname='init',sort = sort)

            else:
                for id in button[key]:
                    test_dict = {}
                    test_dict['id'] = sheet.cell(id+1, 1).value
                    test_dict['module'] = sheet.cell(id+1, 2).value
                    test_dict['title'] = sheet.cell(id+1, 3).value
                    test_dict['method'] = sheet.cell(id+1, 4).value
                    test_dict['url'] = sheet.cell(id+1, 5).value
                    # if sheet.cell(id+1,6).value.find('${sort}')!=-1:
                    #     test_dict['data'] = sheet.cell(id+1,6).value.replace('${sort}',str(sort))
                    #     sort+=1
                    # else:
                    test_dict['data'] = sheet.cell(id + 1, 6).value
                    test_dict['expected'] = sheet.cell(id+1, 7).value
                    test_dict['sheet_name'] = key
                    test_list.append(test_dict)
                    # cls.update_data(sheetname='init', sort = sort)#读取完数据更新数据

        return test_list

    @staticmethod
    def wirt_back(sheet_name,i,result,testresult):
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        sheet.cell(i,8).value = result
        sheet.cell(i,9).value = testresult
        wb.save(file_path)
        '''写回测试数据'''

    @classmethod
    def update_data(cls,sheetname,sort):
        wb = load_workbook(file_path)#定位表
        sheet = wb[sheetname]#订单表单
        sheet.cell(2,1).value = sort#给表格赋值
        wb.save(file_path)#保存
        '''更新对应表格数据'''

if __name__ == '__main__':
    print(DoExcel.read_excel())
